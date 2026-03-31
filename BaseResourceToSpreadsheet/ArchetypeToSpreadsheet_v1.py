"""
ArchetypeToSpreadsheet
Version: 1.0
Author: Wouter Zanen

Purpose
-------
Convert openEHR archetypes from CKM XML into a mapping-friendly Excel overview.

Design goals
------------
- keep the output focused on logical mapping, not full RM detail
- hide routine RM internals where possible
- preserve enough structure to understand events, clusters, slots and coded options
- keep the script as a single readable file with clear progression from download -> parse -> postprocess -> write

Typical flow
------------
1. Resolve an archetype query against CKM
2. Download the XML definition
3. Parse the archetype into mapping rows
4. Post-process rows for readability and deduplication
5. Write the result to Excel
"""

__version__ = "1.0"
__author__ = "Wouter Zanen"

import argparse
import copy
import pathlib
import re
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
import pandas as pd
import requests
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Logging and filesystem safety helpers
# ---------------------------------------------------------------------------
class OutputFileLockedError(RuntimeError):
    """Raised when an output file is open or locked by another process."""


def debug(message: str) -> None:
    print(f"[DEBUG] {message}")


def debug_warning(message: str) -> None:
    print(f"[DEBUG] WARNING: {message}")


def debug_error(message: str) -> None:
    print(f"[DEBUG] ERROR: {message}")


def ensure_output_path_writable(path: pathlib.Path, label: str) -> None:
    if not path.exists():
        return

    try:
        with path.open("r+b"):
            pass
    except PermissionError as exc:
        debug_error(f"{label} is open or locked: {path}")
        raise OutputFileLockedError(
            f"{label} is open or locked: {path}. Close the file and try again."
        ) from exc


# ---------------------------------------------------------------------------
# Conversion orchestration layer
# ---------------------------------------------------------------------------
class ArchetypeToSpreadsheet:
    """Coordinate CKM lookup, XML download, parsing and Excel writing."""
    SUPPRESSED_RM_PATHS = {"language", "encoding", "guideline_id", "workflow_id"}

    COLUMNS = [
        "Element",
        "Aliases",
        "Card",
        "Type",
        "Binding",
        "Definition",
        "Requirements",
        "Dekking",
        "Aanvullende informatie",
    ]

    CKM_BASE = "https://ckm.openehr.org/ckm/rest/v1"

    HIDDEN_RM_FIELDS = {
        "OBSERVATION": [
            ("language", "CODE_PHRASE", "1..1", "Language of the content", "source=RM; path=language"),
            ("encoding", "CODE_PHRASE", "1..1", "Character encoding", "source=RM; path=encoding"),
            ("subject", "PARTY_PROXY", "1..1", "Subject of care (patient or other)", "source=RM; path=subject"),
            ("provider", "PARTY_PROXY", "0..1", "Information provider", "source=RM; path=provider"),
            ("other_participations", "List<PARTICIPATION>", "0..*", "Other participations.", "source=RM; path=other_participations"),
            ("workflow_id", "OBJECT_REF", "0..1", "External workflow reference", "source=RM; path=workflow_id"),
            ("guideline_id", "OBJECT_REF", "0..1", "Guideline reference", "source=RM; path=guideline_id"),
            ("data/origin", "DV_DATE_TIME", "1..1", "Start of the observation history time series.", "source=RM; path=data/origin"),
        ],
        "EVALUATION": [
            ("language", "CODE_PHRASE", "1..1", "Language of the content", "source=RM; path=language"),
            ("encoding", "CODE_PHRASE", "1..1", "Character encoding", "source=RM; path=encoding"),
            ("subject", "PARTY_PROXY", "1..1", "Subject of care", "source=RM; path=subject"),
            ("provider", "PARTY_PROXY", "0..1", "Information provider", "source=RM; path=provider"),
            ("other_participations", "List<PARTICIPATION>", "0..*", "Other participations.", "source=RM; path=other_participations"),
            ("workflow_id", "OBJECT_REF", "0..1", "Workflow reference", "source=RM; path=workflow_id"),
            ("guideline_id", "OBJECT_REF", "0..1", "Guideline reference", "source=RM; path=guideline_id"),
        ],
        "INSTRUCTION": [
            ("language", "CODE_PHRASE", "1..1", "Language", "source=RM; path=language"),
            ("encoding", "CODE_PHRASE", "1..1", "Character encoding", "source=RM; path=encoding"),
            ("subject", "PARTY_PROXY", "1..1", "Subject of care", "source=RM; path=subject"),
            ("provider", "PARTY_PROXY", "0..1", "Information provider", "source=RM; path=provider"),
            ("other_participations", "List<PARTICIPATION>", "0..*", "Other participations.", "source=RM; path=other_participations"),
            ("workflow_id", "OBJECT_REF", "0..1", "Workflow reference", "source=RM; path=workflow_id"),
            ("guideline_id", "OBJECT_REF", "0..1", "Guideline reference", "source=RM; path=guideline_id"),
            ("expiry_time", "DV_DATE_TIME", "0..1", "Expiry time of the instruction.", "source=RM; path=expiry_time"),
        ],
        "ACTION": [
            ("language", "CODE_PHRASE", "1..1", "Language", "source=RM; path=language"),
            ("encoding", "CODE_PHRASE", "1..1", "Character encoding", "source=RM; path=encoding"),
            ("subject", "PARTY_PROXY", "1..1", "Subject of care", "source=RM; path=subject"),
            ("provider", "PARTY_PROXY", "0..1", "Information provider", "source=RM; path=provider"),
            ("other_participations", "List<PARTICIPATION>", "0..*", "Other participations.", "source=RM; path=other_participations"),
            ("workflow_id", "OBJECT_REF", "0..1", "Workflow reference", "source=RM; path=workflow_id"),
            ("guideline_id", "OBJECT_REF", "0..1", "Guideline reference", "source=RM; path=guideline_id"),
            ("time", "DV_DATE_TIME", "1..1", "Time of the action.", "source=RM; path=time"),
            ("ism_transition", "ISM_TRANSITION", "1..1", "Workflow state transition.", "source=RM; path=ism_transition"),
            ("instruction_details", "INSTRUCTION_DETAILS", "0..1", "Link to instruction.", "source=RM; path=instruction_details"),
        ],
        "ADMIN_ENTRY": [
            ("language", "CODE_PHRASE", "1..1", "Language", "source=RM; path=language"),
            ("encoding", "CODE_PHRASE", "1..1", "Character encoding", "source=RM; path=encoding"),
            ("subject", "PARTY_PROXY", "1..1", "Subject", "source=RM; path=subject"),
            ("provider", "PARTY_PROXY", "0..1", "Information provider", "source=RM; path=provider"),
            ("other_participations", "List<PARTICIPATION>", "0..*", "Other participations.", "source=RM; path=other_participations"),
            ("workflow_id", "OBJECT_REF", "0..1", "Workflow reference", "source=RM; path=workflow_id"),
        ],
    }

    def __init__(
        self,
        output_folder: str,
        language: str = "en",
        username: Optional[str] = None,
        password: Optional[str] = None,
        session_id: Optional[str] = None,
        timeout: int = 60,
    ):
        self.output_folder = pathlib.Path(output_folder)
        self.output_folder.mkdir(parents=True, exist_ok=True)
        self.language = language
        self.timeout = timeout
        self.session = requests.Session()
        if username and password:
            self.session.auth = (username, password)
        if session_id:
            self.session.cookies.set("JSESSIONID", session_id)

    def convert(self, archetype_query: str, published_only: bool = True, fetch_mode: str = "auto") -> pathlib.Path:
        """Run the full conversion pipeline for a single archetype query."""
        request = self._resolve_conversion_request(archetype_query)
        xml_text, _resolved_fetch_mode = self._download_conversion_xml(
            request,
            published_only=published_only,
            fetch_mode=fetch_mode,
        )
        # 3. Persist the raw XML so parser/debug issues can be inspected later.
        self._write_debug_xml(xml_text)

        # 4. Parse the archetype into logical mapping rows.
        parser, rows = self._parse_archetype_rows(xml_text)
        final_archetype_id = self._finalize_archetype_id(
            parser,
            request["requested_archetype_id"],
        )

        entry_type = self._infer_entry_type(final_archetype_id, parser.archetype_id)
        # 5. Apply row-level cleanup such as hidden RM rows, dedupe and sorting.
        rows = self._postprocess_rows(rows, entry_type)
        debug(f"Total rows (after RM): {len(rows)}")

        writer = WorkbookWriter(self.COLUMNS)
        return writer.write_rows(
            rows=rows,
            output_folder=self.output_folder,
            base_name=self._derive_base_name(final_archetype_id),
        )

    def _resolve_conversion_request(self, archetype_query: str) -> Dict[str, object]:
        meta = self._find_archetype(archetype_query)
        cid = str(meta.get("citeableId") or meta.get("citeableIdentifier") or meta.get("cid") or meta.get("id"))
        requested_archetype_id = str(
            meta.get("archetypeId") or meta.get("resourceMainId") or meta.get("resourceId") or archetype_query
        )
        return {
            "meta": meta,
            "cid": cid,
            "requested_archetype_id": requested_archetype_id,
        }

    def _download_conversion_xml(
        self,
        request: Dict[str, object],
        published_only: bool = True,
        fetch_mode: str = "auto",
    ) -> Tuple[str, str]:
        requested_archetype_id = str(request["requested_archetype_id"])
        cid = str(request["cid"])
        resolved_fetch_mode = self._resolve_fetch_mode(requested_archetype_id, fetch_mode)
        xml_text = self._download_archetype_xml(
            cid,
            published_only=published_only,
            fetch_mode=resolved_fetch_mode,
        )
        debug(f"Requested archetype ID: {requested_archetype_id}")
        debug(f"Fetch mode: {resolved_fetch_mode}")
        debug(f"CID: {cid}")
        debug(f"XML length: {len(xml_text)}")
        return xml_text, resolved_fetch_mode

    def _write_debug_xml(self, xml_text: str) -> pathlib.Path:
        debug_xml = self.output_folder / "debug_archetype.xml"
        ensure_output_path_writable(debug_xml, "Debug XML output file")
        debug_xml.write_text(xml_text, encoding="utf-8")
        debug(f"Saved XML to: {debug_xml}")
        return debug_xml

    def _parse_archetype_rows(self, xml_text: str) -> Tuple["ArchetypeXmlParser", List["MappingRow"]]:
        parser = ArchetypeXmlParser(xml_text, preferred_language=self.language)
        rows = parser.to_rows()
        for warning in parser.debug_warnings:
            debug_warning(warning)
        debug(f"Parsed rows (before RM): {len(rows)}")
        return parser, rows

    def _finalize_archetype_id(self, parser: "ArchetypeXmlParser", requested_archetype_id: str) -> str:
        if parser.archetype_id and parser.archetype_id.lower().startswith("openehr-"):
            return parser.archetype_id
        return requested_archetype_id

    def _postprocess_rows(self, rows: List["MappingRow"], entry_type: str) -> List["MappingRow"]:
        rows = self._prepend_hidden_rm_rows(rows, entry_type)
        rows = self._merge_duplicate_elements(rows)
        rows = self._sort_rows(rows)
        return rows

    def _derive_base_name(self, archetype_id: str) -> str:
        match = re.match(r"^openEHR-[A-Z]+-[A-Z_]+\.(.+?)\.v\d+$", archetype_id, flags=re.IGNORECASE)
        if match:
            name = match.group(1)
        else:
            name = archetype_id
        name = re.sub(r"[^A-Za-z0-9_.-]+", "_", name)
        name = name.strip("._-")
        return name or "archetype"

    def _safe_sheet_name(self, name: str) -> str:
        cleaned = re.sub(r"[:\\/*?\[\]]", "_", name)
        return cleaned[:31] or "Sheet1"

    def _normalize_archetype_query(self, archetype_query: str) -> str:
        normalized = (archetype_query or "").strip()
        if normalized.lower().endswith(".xml"):
            normalized = normalized[:-4]
        return normalized

    def _find_archetype(self, archetype_query: str) -> dict:
        """Resolve an archetype id or search query to CKM metadata."""
        normalized_query = self._normalize_archetype_query(archetype_query)

        if normalized_query.startswith("openEHR-"):
            cid = self._get_citeable_identifier(normalized_query)
            if cid:
                return {"cid": cid, "archetypeId": normalized_query}

        search_candidates = [normalized_query]
        if normalized_query.startswith("openEHR-EHR-"):
            search_candidates.append(normalized_query.replace("openEHR-EHR-", "", 1))
            parts = normalized_query.split(".")
            if len(parts) >= 2 and parts[1]:
                search_candidates.append(parts[1])

        seen_candidates = set()
        for candidate in search_candidates:
            lowered = candidate.lower()
            if lowered in seen_candidates:
                continue
            seen_candidates.add(lowered)

            params = {
                "search-text": candidate,
                "restrict-search-to-main-data": "true",
                "require-all-search-words": "false",
                "language": self.language,
                "size": 50,
                "offset": 0,
            }
            response = self.session.get(
                f"{self.CKM_BASE}/archetypes",
                params=params,
                headers={"Accept": "application/json"},
                timeout=self.timeout,
            )
            response.raise_for_status()
            payload = self._safe_json(response, endpoint="/archetypes search")
            items = payload if isinstance(payload, list) else payload.get("items") or payload.get("resources") or payload.get("result") or []
            if not items:
                continue

            def item_id(item: dict) -> str:
                return str(item.get("archetypeId") or item.get("resourceMainId") or item.get("resourceId") or "").strip()

            exact_matches = [item for item in items if item_id(item).lower() == normalized_query.lower()]
            partial_matches = [
                item for item in items
                if normalized_query.lower() in item_id(item).lower()
                or candidate.lower() in item_id(item).lower()
            ]
            chosen = exact_matches[0] if exact_matches else partial_matches[0] if partial_matches else items[0]

            if "citeableId" not in chosen and "citeableIdentifier" not in chosen and "cid" not in chosen and "id" not in chosen:
                resource_id = item_id(chosen)
                cid = self._get_citeable_identifier(resource_id)
                if cid:
                    chosen["cid"] = cid
            if "citeableId" in chosen or "citeableIdentifier" in chosen or "cid" in chosen or "id" in chosen:
                return chosen

        raise ValueError(f"No CKM archetype found for query: {archetype_query}")

    def _get_citeable_identifier(self, archetype_id: str) -> Optional[str]:
        response = self.session.get(
            f"{self.CKM_BASE}/archetypes/citeable-identifier/{archetype_id}",
            headers={"Accept": "text/plain"},
            timeout=self.timeout,
        )
        if response.status_code != 200:
            return None
        cid = response.text.strip()
        return cid or None

    def _safe_json(self, response: requests.Response, endpoint: str):
        try:
            return response.json()
        except Exception as e:
            snippet = (response.text or "")[:300].replace("\n", " ").replace("\r", " ")
            content_type = response.headers.get("Content-Type", "")
            raise ValueError(
                f"CKM endpoint {endpoint} did not return JSON. "
                f"Status={response.status_code}, Content-Type={content_type!r}, Body starts with: {snippet!r}"
            ) from e

    def _major_version_from_archetype_id(self, archetype_id: str) -> Optional[int]:
        match = re.search(r"\.v(\d+)$", (archetype_id or "").strip())
        return int(match.group(1)) if match else None

    def _resolve_fetch_mode(self, archetype_id: str, fetch_mode: str) -> str:
        normalized_mode = (fetch_mode or "auto").strip().lower()
        if normalized_mode not in {"auto", "published", "development"}:
            raise ValueError(f"Unsupported fetch mode: {fetch_mode}")

        if normalized_mode != "auto":
            return normalized_mode

        major = self._major_version_from_archetype_id(archetype_id)
        if major is not None and major == 0:
            return "development"
        return "published"

    def _download_archetype_xml(self, cid: str, published_only: bool = True, fetch_mode: Optional[str] = None) -> str:
        if fetch_mode is not None:
            mode = fetch_mode.strip().lower()
            if mode not in {"published", "development"}:
                raise ValueError(f"Unsupported fetch mode: {fetch_mode}")
            published_only = (mode == "published")

        params = {}
        if published_only:
            params["get-latest-published"] = "true"
        response = self.session.get(f"{self.CKM_BASE}/archetypes/{cid}/xml", params=params, timeout=self.timeout)
        if response.status_code == 404 and published_only:
            response = self.session.get(f"{self.CKM_BASE}/archetypes/{cid}/xml", timeout=self.timeout)
        response.raise_for_status()
        return response.text

    def _infer_entry_type(self, *candidates: Optional[str]) -> Optional[str]:
        for candidate in candidates:
            if not candidate:
                continue
            upper = candidate.upper()
            for entry_type in self.HIDDEN_RM_FIELDS:
                if entry_type in upper:
                    return entry_type
        return None

    def _prepend_hidden_rm_rows(self, rows: List[MappingRow], entry_type: Optional[str]) -> List[MappingRow]:
        if not entry_type:
            return rows
        hidden_rows: List[MappingRow] = []
        for path, typ, card, definition, info in self.HIDDEN_RM_FIELDS[entry_type]:
            if path in self.SUPPRESSED_RM_PATHS:
                continue
            hidden_rows.append(
                MappingRow(
                    element=self._display_rm_element(path),
                    card=card,
                    datatype=typ,
                    definition=definition,
                    info=info,
                )
            )
        existing_elements = {row.element for row in rows}
        return [row for row in hidden_rows if row.element not in existing_elements] + rows

    def _display_rm_element(self, path: str) -> str:
        label = path.split("/")[-1].replace("_", " ").strip().capitalize()
        if path == "data/origin":
            label = "Origin"

        def norm(s: str) -> str:
            return re.sub(r"[^a-z0-9]+", "", s.lower())

        # Only show the RM path when it adds information.
        # Examples:
        # - "Encoding (encoding)" -> redundant, so show just "Encoding"
        # - "Origin (data/origin)" -> informative, so keep the path
        if "/" not in path and norm(label) == norm(path):
            return label
        return f"{label} ({path})"

    def _parse_info(self, info: str) -> Dict[str, str]:
        parsed: Dict[str, str] = {}
        for part in (info or "").split(";"):
            if "=" not in part:
                continue
            key, value = part.split("=", 1)
            key = key.strip()
            value = value.strip()
            if key:
                parsed[key] = value
        return parsed

    def _merge_key(self, row: MappingRow) -> Tuple[str, str, str]:
        info = self._parse_info(row.info)
        return (
            row.element,
            info.get("path", ""),
            info.get("node_id", ""),
        )

    def _merge_duplicate_elements(self, rows: List[MappingRow]) -> List[MappingRow]:
        grouped: Dict[Tuple[str, str, str], List[MappingRow]] = {}
        order: List[Tuple[str, str, str]] = []
        for row in rows:
            key = self._merge_key(row)
            if key not in grouped:
                grouped[key] = []
                order.append(key)
            grouped[key].append(row)

        merged_rows: List[MappingRow] = []
        for key in order:
            entries = grouped[key]
            element = entries[0].element

            def collect_multi(field_name: str, sort_values: bool = False) -> str:
                values: List[str] = []
                seen = set()
                for entry in entries:
                    raw = str(getattr(entry, field_name) or "").strip()
                    if not raw:
                        continue
                    parts = [p.strip() for p in raw.split(",")]
                    for part in parts:
                        if part and part not in seen:
                            seen.add(part)
                            values.append(part)
                if sort_values:
                    values = sorted(values, key=lambda x: (0 if x == "DV_CODED_TEXT" else 1 if x == "DV_TEXT" else 2, x))
                return ", ".join(values)

            def pick_single(field_name: str) -> str:
                values = [str(getattr(entry, field_name) or "").strip() for entry in entries if str(getattr(entry, field_name) or "").strip()]
                if not values:
                    return ""
                if field_name == "definition":
                    return max(values, key=len)
                return values[0]

            merged_rows.append(
                MappingRow(
                    element=element,
                    card=pick_single("card"),
                    datatype=collect_multi("datatype", sort_values=True),
                    binding=collect_multi("binding"),
                    definition=pick_single("definition"),
                    coverage=collect_multi("coverage"),
                    info=collect_multi("info"),
                )
            )
        return merged_rows

    def _sort_rows(self, rows: List[MappingRow]) -> List[MappingRow]:
        # Preserve traversal order so cluster rows stay directly above their children.
        # Hidden RM rows are already prepended before archetype rows in convert().
        return rows

    def _row_depth_from_info(self, info: str) -> int:
        return WorkbookWriter.row_depth_from_info(info)

    def _style_workbook(self, workbook_path: pathlib.Path, sheet_name: str) -> None:
        writer = WorkbookWriter(self.COLUMNS)
        writer.style_workbook(workbook_path, sheet_name)




# ---------------------------------------------------------------------------
# Workbook writing and styling
# ---------------------------------------------------------------------------
class WorkbookWriter:
    """Write MappingRow instances to Excel and apply lightweight formatting."""
    def __init__(self, columns: List[str]):
        self.columns = columns

    def write_rows(self, rows: List["MappingRow"], output_folder: pathlib.Path, base_name: str) -> pathlib.Path:
        dataframe = self._build_dataframe(rows)
        workbook_path = output_folder / f"{base_name}_base.xlsx"
        sheet_name = self.safe_sheet_name(f"{base_name}_base")

        ensure_output_path_writable(workbook_path, "Excel output file")
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

        self.style_workbook(workbook_path, sheet_name)
        return workbook_path

    def _build_dataframe(self, rows: List["MappingRow"]) -> pd.DataFrame:
        return pd.DataFrame([row.to_list() for row in rows], columns=self.columns)

    @staticmethod
    def safe_sheet_name(name: str) -> str:
        cleaned = re.sub(r"[:\\/*?\[\]]", "_", name)
        return cleaned[:31] or "Sheet1"

    @staticmethod
    def row_depth_from_info(info: str) -> int:
        match = re.search(r"(?:^|;)\s*depth=(\d+)", str(info or ""))
        return int(match.group(1)) if match else 0

    def style_workbook(self, workbook_path: pathlib.Path, sheet_name: str) -> None:
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook[sheet_name]

        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(vertical="top")

        info_col_idx = len(self.columns)
        for row_idx in range(2, sheet.max_row + 1):
            info_value = sheet.cell(row=row_idx, column=info_col_idx).value
            depth = self.row_depth_from_info(info_value)
            name_cell = sheet.cell(row=row_idx, column=1)
            name_cell.alignment = openpyxl.styles.Alignment(
                indent=min(depth, 15),
                vertical="top",
                wrap_text=True,
            )
            for col_idx in range(2, sheet.max_column + 1):
                sheet.cell(row=row_idx, column=col_idx).alignment = openpyxl.styles.Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in sheet.columns:
            cleaned_lengths = []
            for cell in column_cells:
                if cell.value is None:
                    cleaned_lengths.append(0)
                else:
                    cleaned_lengths.append(len(str(cell.value).strip()))
            max_length = max(cleaned_lengths) if cleaned_lengths else 0
            extra = 4 if column_cells[0].column == 1 else 2
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + extra, 12), 90)

        ensure_output_path_writable(workbook_path, "Excel output file")
        workbook.save(workbook_path)


@dataclass
# ---------------------------------------------------------------------------
# Lightweight data model used by the parser
# ---------------------------------------------------------------------------
class ParsedTerm:
    """Human-readable ontology text/comment/description for a node id."""
    text: str = ""
    description: str = ""
    comment: str = ""


@dataclass
class MappingRow:
    """One logical output row in the mapping spreadsheet."""
    element: str
    aliases: str = ""
    card: str = ""
    datatype: str = ""
    binding: str = ""
    definition: str = ""
    requirements: str = ""
    coverage: str = ""
    info: str = ""

    def to_list(self) -> List[str]:
        return [
            self.element,
            self.aliases,
            self.card,
            self.datatype,
            self.binding,
            self.definition,
            self.requirements,
            self.coverage,
            self.info,
        ]


# ---------------------------------------------------------------------------
# XML parser and display policy
# ---------------------------------------------------------------------------
class ArchetypeXmlParser:
    """Parse archetype XML into readable, mapping-oriented rows."""
    OPENEHR_TERMINOLOGY_URL = "https://specifications.openehr.org/releases/TERM/development/computable/XML/en/openehr_terminology.xml"
    _OPENEHR_TERMINOLOGY_CACHE: Optional[Dict[str, str]] = None
    _OPENEHR_TERMINOLOGY_ERROR: Optional[str] = None
    KEEP_RM_TYPES = {"CLUSTER"}
    EVENT_TYPES = {"EVENT", "POINT_EVENT", "INTERVAL_EVENT"}

    def __init__(self, xml_text: str, preferred_language: str = "en"):
        self.root = ET.fromstring(xml_text)
        self.preferred_language = preferred_language
        self.terms = self._extract_terms()
        self.bindings = self._extract_bindings()
        self.archetype_id = self._extract_archetype_id()
        self.definition = self._find_first(self.root, "definition")
        self.path_index = self._build_object_path_index(self.definition)
        self.debug_warnings: List[str] = []

    def _object_rm_type(self, element: ET.Element) -> str:
        return self._rm_type_name(element) or self._primitive_type_name(element) or self._xsi_type(element) or "UNKNOWN"

    def _object_node_id(self, element: ET.Element) -> str:
        return self._text_of_first(element, "node_id") or element.attrib.get("node_id") or ""

    def _attribute_name(self, attr: ET.Element) -> str:
        return self._text_of_child(attr, "rm_attribute_name") or self._text_of_child(attr, "name") or "attribute"

    def _build_archetype_info(
        self,
        visible_path: str,
        depth: int,
        child_code: str,
        target_path: Optional[str],
        actual_xsi_type: Optional[str],
        is_internal_ref: bool,
        raw_rm_type: str,
        effective_rm_type: str,
    ) -> List[str]:
        info = ["source=Archetype", f"path={visible_path}", f"depth={depth}"]
        if child_code:
            info.append(f"node_id={child_code}")
        if target_path:
            info.append(f"target_path={self._visible_path(target_path)}")
        if actual_xsi_type:
            info.append(f"xsi_type={actual_xsi_type}")
        if is_internal_ref and not actual_xsi_type:
            info.append("internal_ref=true")
        elif raw_rm_type and raw_rm_type != effective_rm_type:
            info.append(f"proxy_type={raw_rm_type}")
        return info

    def _should_skip_event_rm_leaf(self, event_label: Optional[str], attr_name: str) -> bool:
        return bool(event_label and attr_name in {"time", "math_function", "width"})

    def _is_ism_transition_path(self, parent_path: str, path: str) -> bool:
        normalized_parent = re.sub(r"\[[^\]]+\]", "", parent_path or "")
        normalized_path = re.sub(r"\[[^\]]+\]", "", path or "")
        return "ism_transition" in normalized_parent or "ism_transition" in normalized_path

    def _should_skip_ism_transition_leaf(self, parent_path: str, path: str, attr_name: str) -> bool:
        return self._is_ism_transition_path(parent_path, path) and attr_name in {"current_state", "careflow_step", "transition"}

    def _extract_attribute_code(
        self,
        obj: ET.Element,
        attr_name: str,
        terminology_hint: Optional[str] = None,
    ) -> str:
        target = self._find_attr_child(obj, attr_name)
        if target is None:
            return ""
        return self._extract_code_string(target, terminology_hint=terminology_hint)

    def _extract_attribute_display(
        self,
        obj: ET.Element,
        attr_name: str,
        terminology_hint: Optional[str] = None,
    ) -> str:
        code = self._extract_attribute_code(obj, attr_name, terminology_hint=terminology_hint)
        if not code:
            return ""
        terminology_id = terminology_hint or "local"
        return self._resolve_coded_option_label(code, terminology_id)

    def to_rows(self) -> List[MappingRow]:
        """Parse the archetype definition into the final logical row list."""
        definition = self.definition
        if definition is None:
            raise ValueError("Could not locate archetype definition in CKM XML.")

        rows = self._walk_object(
            definition,
            parent_path="",
            label_term=self.terms.get(self._text_of_first(definition, "node_id") or "", ParsedTerm()),
            label_code=self._text_of_first(definition, "node_id") or "",
            event_label=None,
        )
        rows.extend(self._extract_ism_transition_variant_rows(definition))
        return self._dedupe_rows(rows)

    def _extract_ism_transition_variant_rows(self, definition: ET.Element) -> List[MappingRow]:
        """Flatten ISM_TRANSITION variants into one row per constrained careflow step."""
        rows: List[MappingRow] = []
        # ISM_TRANSITION is modelled as a compact set of constraints in the XML.
        # For mapping purposes we expose each constrained state as a standalone row.
        for obj in definition.iter():
            rm_type = self._rm_type_name(obj) or self._primitive_type_name(obj) or self._xsi_type(obj) or ""
            node_id = self._text_of_first(obj, "node_id") or obj.attrib.get("node_id") or ""
            if rm_type != "ISM_TRANSITION" or not node_id:
                continue

            term = self.terms.get(node_id, ParsedTerm())
            base_label = term.text or node_id
            careflow_display = self._extract_ism_careflow_step_display(obj) or ""
            visible_label = careflow_display or base_label

            # Deprecated variants should be skipped before any merge/deduping.
            if not visible_label or visible_label.lower().startswith("x - ") or base_label.lower().startswith("x - "):
                continue

            definition_text = self._definition_text(term)
            if self._is_internal_definition(definition_text):
                definition_text = ""

            current_state_code = self._extract_ism_current_state_code(obj)
            current_state_display = self._extract_ism_current_state_display(obj)

            extras = []
            if careflow_display:
                extras.append(f"Careflow step: {careflow_display}.")
            if current_state_display and current_state_code:
                extras.append(f"Current state: {current_state_display} (openehr::{current_state_code}).")
            elif current_state_display:
                extras.append(f"Current state: {current_state_display}.")

            full_def = definition_text
            if extras:
                full_def = self._append_sentence(full_def, " ".join(extras))

            rows.append(
                MappingRow(
                    element=f"{visible_label} (ism_transition/careflow_step)",
                    card="conditional",
                    datatype="DV_CODED_TEXT",
                    definition=full_def.strip(),
                    info=f"source=Archetype; node_id={node_id}; path=ism_transition[{node_id}]/careflow_step",
                )
            )
        return rows

    def _extract_ism_current_state_code(self, ism_obj: ET.Element) -> str:
        return self._extract_attribute_code(ism_obj, "current_state", terminology_hint="openehr")

    def _extract_ism_current_state_display(self, ism_obj: ET.Element) -> str:
        return self._extract_attribute_display(ism_obj, "current_state", terminology_hint="openehr")

    def _extract_ism_careflow_step_display(self, ism_obj: ET.Element) -> str:
        return self._extract_attribute_display(ism_obj, "careflow_step", terminology_hint="local")

    def _find_attr_child(self, obj: ET.Element, attr_name: str) -> Optional[ET.Element]:
        # Then inspect each constrained RM attribute and decide whether it becomes
        # an event row, cluster row, DV leaf row or a recursion-only container.
        for attr in self._children_named(obj, "attributes"):
            name = self._text_of_child(attr, "rm_attribute_name") or self._text_of_child(attr, "name") or ""
            if name != attr_name:
                continue
            children = self._child_constraint_objects(attr)
            if children:
                return children[0]
        return None

    def _extract_code_string(self, element: Optional[ET.Element], terminology_hint: Optional[str] = None) -> str:
        """Extract a code value from CKM XML, supporting both raw code_list and [term::code] shapes."""
        if element is None:
            return ""

        normalized_hint = (terminology_hint or "").strip().lower()

        # First try the common CKM XML structure:
        # terminology_id/value -> code_list
        if normalized_hint:
            seen_matching_terminology = False
            for elem in element.iter():
                local = self._local(elem.tag)
                value = (elem.text or "").strip()
                if not value:
                    continue

                if local == "value" and value.lower() == normalized_hint:
                    seen_matching_terminology = True
                    continue

                if local in {"code_list", "code_string", "code"} and seen_matching_terminology:
                    return value

        # Next, support explicit bracket notation such as [openehr::146].
        candidates = []
        for elem in element.iter():
            value = (elem.text or "").strip()
            if value:
                candidates.append(value)
            for attr_val in elem.attrib.values():
                if isinstance(attr_val, str) and attr_val:
                    candidates.append(attr_val)

        pattern = r"\[(?:(?P<term>[A-Za-z0-9_-]+)::)?(?P<code>[A-Za-z0-9_.-]+)\]"
        for candidate in candidates:
            match = re.search(pattern, candidate)
            if not match:
                continue
            term = (match.group("term") or "").lower()
            code = match.group("code")
            if normalized_hint and term and term != normalized_hint:
                continue
            return code

        # Finally, fall back to the first raw code_list/code value when no terminology
        # branch could be identified. This keeps local/openEHR coded text extraction working
        # for compact CKM XML shapes.
        for elem in element.iter():
            local = self._local(elem.tag)
            value = (elem.text or "").strip()
            if local in {"code_list", "code_string", "code"} and value:
                return value

        return ""

    def _append_sentence(self, base: str, extra: str) -> str:
        base = (base or "").strip()
        extra = (extra or "").strip()
        if not base:
            return extra
        if not extra:
            return base
        if extra.lower() in base.lower():
            return base
        if not base.endswith("."):
            base += "."
        return base + " " + extra

    def _build_object_path_index(self, definition: Optional[ET.Element]) -> Dict[str, ET.Element]:
        index: Dict[str, ET.Element] = {}
        if definition is None:
            return index
        self._index_object_paths(definition, "", index)
        return index

    def _index_object_paths(self, obj: ET.Element, parent_path: str, index: Dict[str, ET.Element]) -> None:
        visible_parent_path = self._visible_path(parent_path)
        if visible_parent_path:
            index[self._normalize_constraint_path(visible_parent_path)] = obj

        for attr in self._children_named(obj, "attributes"):
            attr_name = self._text_of_child(attr, "rm_attribute_name") or self._text_of_child(attr, "name") or "attribute"
            for child in self._child_constraint_objects(attr):
                node_id = self._object_node_id(child)
                child_path = self._build_path(parent_path, attr_name, node_id or None)
                self._index_object_paths(child, child_path, index)

    def _normalize_constraint_path(self, path: str) -> str:
        return re.sub(r"/+", "/", (path or "").strip()).strip("/")

    def _resolve_reference_target(self, element: ET.Element) -> Optional[ET.Element]:
        target_path = self._text_of_child(element, "target_path") or ""
        if not target_path:
            return None
        normalized = self._normalize_constraint_path(target_path)
        return self.path_index.get(normalized)

    def _effective_rm_type(self, element: ET.Element, target_obj: Optional[ET.Element] = None) -> str:
        target = target_obj
        if target is None and self._is_internal_ref(element):
            target = self._resolve_reference_target(element)
        if target is not None:
            effective = self._rm_type_name(target) or self._primitive_type_name(target) or self._xsi_type(target)
            if effective:
                return effective
        return self._rm_type_name(element) or self._primitive_type_name(element) or self._xsi_type(element) or "UNKNOWN"

    def _is_internal_ref(self, element: ET.Element) -> bool:
        xsi = (self._xsi_type(element) or "").upper()
        if xsi in {"ARCHETYPE_INTERNAL_REF", "ARCHETYPE_INTERNAL_REFERENCE", "C_COMPLEX_OBJECT_PROXY"}:
            return True
        target_path = self._text_of_first(element, "target_path") or ""
        has_attributes = bool(self._children_named(element, "attributes"))
        has_includes = bool(self._children_named(element, "includes"))
        rm_type = (self._rm_type_name(element) or "").upper()
        return bool(target_path) and not has_attributes and not has_includes and rm_type in {"", "CLUSTER"}

    def _indent_name(self, text: str, depth: int) -> str:
        return (text or "").strip()

    def _display_event(self, term: ParsedTerm, path: str) -> str:
        label = term.text or self._prettify_segment(path.split("/")[-1])
        return f"{label} (events)" if label else "events"

    def _walk_object(
        self,
        obj: ET.Element,
        parent_path: str,
        label_term: ParsedTerm,
        label_code: str,
        event_label: Optional[str],
        internal_cluster_context: Optional[str] = None,
        depth: int = 0,
    ) -> List[MappingRow]:
        """Recursively walk one constrained object and emit mapping rows."""
        rows: List[MappingRow] = []
        obj_rm_type = self._object_rm_type(obj)
        obj_occ_card = self._cardinality_from_object(obj)

        # When recursion enters an event object, first emit its RM-level event fields
        # (for example time, math function or width) before descending into its payload.
        if event_label and obj_rm_type in self.EVENT_TYPES:
            rows.extend(self._event_rm_rows(event_label, obj_rm_type, obj, depth=depth))

        for attr in self._children_named(obj, "attributes"):
            attr_name = self._attribute_name(attr)
            attr_card = self._cardinality_from_attribute(attr) or ""
            child_objects = self._child_constraint_objects(attr)

            for child in child_objects:
                node_id = self._text_of_first(child, "node_id") or child.attrib.get("node_id") or ""
                child_term = self.terms.get(node_id, label_term) if node_id else label_term
                child_code = node_id or label_code

                path = self._build_path(parent_path, attr_name, node_id or None)
                raw_rm_type = self._object_rm_type(child)
                is_slot = self._is_cluster_slot(child)
                is_internal_ref = self._is_internal_ref(child)

                target_obj = None
                if is_internal_ref:
                    candidate = self._resolve_reference_target(child)
                    if candidate is not None:
                        target_obj = candidate

                effective_obj = target_obj if target_obj is not None else child
                rm_type = self._effective_rm_type(child, target_obj)
                binding = self.bindings.get(child_code, "")
                target_path = self._text_of_child(child, "target_path")
                definition = self._definition_text(child_term)
                if self._is_internal_definition(definition):
                    definition = ""

                visible_path = self._visible_path(path)
                actual_xsi_type = self._xsi_type(child)
                info = self._build_archetype_info(
                    visible_path=visible_path,
                    depth=depth,
                    child_code=child_code,
                    target_path=target_path,
                    actual_xsi_type=actual_xsi_type,
                    is_internal_ref=is_internal_ref,
                    raw_rm_type=raw_rm_type,
                    effective_rm_type=rm_type,
                )

                # Explicit event children become visible event rows and increase depth.
                if rm_type in self.EVENT_TYPES and child_term.text:
                    event_display = self._display_event(child_term, path)
                    rows.append(
                        MappingRow(
                            element=self._indent_name(event_display, depth),
                            card=self._cardinality_from_object(child) or attr_card or "0..1",
                            datatype=rm_type,
                            definition=definition,
                            info="; ".join(info),
                        )
                    )
                    rows.extend(self._walk_object(
                        effective_obj,
                        path,
                        child_term,
                        child_code,
                        child_term.text,
                        internal_cluster_context,
                        depth + 1,
                    ))
                    continue

                # Keep visible clusters and slots, but treat slots differently later
                # by not pretending they are internal structure definitions.
                if rm_type in self.KEEP_RM_TYPES or is_slot:
                    cluster_type = "SLOT (CLUSTER)" if is_slot else "CLUSTER"
                    cluster_definition = ((definition or "") + self._cluster_includes_text(child)).strip()
                    cluster_label = child_term.text or self._prettify_segment(path.split("/")[-1])
                    rows.append(
                        MappingRow(
                            element=self._indent_name(self._display_cluster(child_term, path, internal_cluster_context), depth),
                            card=self._concept_cardinality(child, attr, parent_object_occurrences=obj_occ_card),
                            datatype=cluster_type,
                            binding=binding,
                            definition=cluster_definition,
                            info="; ".join(info),
                        )
                    )
                    next_cluster_context = internal_cluster_context if is_slot else cluster_label
                    rows.extend(self._walk_object(
                        effective_obj,
                        path,
                        child_term,
                        child_code,
                        event_label,
                        next_cluster_context,
                        depth + 1,
                    ))
                    continue

                # Primitive DV_* leaves are the main payload rows used for mapping.
                if rm_type.startswith("DV_"):
                    if self._should_skip_event_rm_leaf(event_label, attr_name):
                        continue
                    if self._should_skip_ism_transition_leaf(parent_path, path, attr_name):
                        continue

                    display_element = self._display_leaf(visible_path, child_term, event_label, internal_cluster_context)
                    display_definition = self._enrich_definition(attr_name, definition, event_label)
                    if rm_type == "DV_CODED_TEXT":
                        display_definition = self._append_coded_text_options(display_definition, child, attr_name)
                    leaf_depth = depth + 1 if event_label and attr_name in {"data", "state"} else depth
                    rows.append(
                        MappingRow(
                            element=self._indent_name(display_element, leaf_depth),
                            card=self._concept_cardinality(obj, attr, parent_object_occurrences=obj_occ_card),
                            datatype=rm_type,
                            binding=binding,
                            definition=display_definition,
                            info="; ".join(info),
                        )
                    )
                    continue

                rows.extend(self._walk_object(
                    effective_obj,
                    path,
                    child_term,
                    child_code,
                    event_label,
                    internal_cluster_context,
                    depth,
                ))

        return rows

    def _event_math_function_definition(self, event_obj: ET.Element) -> str:
        math_function_obj = self._find_attr_child(event_obj, "math_function")
        if math_function_obj is None:
            return "Mathematical function used for the interval summary."

        options = self._extract_coded_text_options(math_function_obj)
        if options:
            return f"DV_CODED_TEXT (value in): {'; '.join(options)}"

        code = self._extract_code_string(math_function_obj, terminology_hint="openehr")
        if code:
            resolved = self._resolve_coded_option_label(code, "openehr")
            return f"DV_CODED_TEXT (value in): {resolved}"

        return "Mathematical function used for the interval summary."

    def _event_width_definition(self, event_obj: ET.Element, event_label: str) -> str:
        width_obj = self._find_attr_child(event_obj, "width")
        if width_obj is not None:
            code = self._extract_code_string(width_obj)
            if code:
                return code
        if event_label.lower() == "24 hour average":
            return "PT24H"
        return "Duration of the interval represented by this event."

    def _event_rm_rows(self, event_label: str, rm_type: str, event_obj: ET.Element, depth: int = 0) -> List[MappingRow]:
        """Create the visible RM rows that belong to an event wrapper."""
        rows = []
        if rm_type in {"EVENT", "POINT_EVENT"}:
            rows.append(
                MappingRow(
                    element=f"Time ({event_label})",
                    card="1..1",
                    datatype="DV_DATE_TIME",
                    definition="Time of this event occurrence.",
                    info=f"source=RM; path=data/events/time; event={event_label}; depth={depth}",
                )
            )
        if rm_type == "INTERVAL_EVENT":
            rows.append(
                MappingRow(
                    element=f"Math function ({event_label})",
                    card="1..1",
                    datatype="DV_CODED_TEXT",
                    definition=self._event_math_function_definition(event_obj),
                    info=f"source=RM; path=data/events/math_function; event={event_label}; depth={depth}",
                )
            )
            rows.append(
                MappingRow(
                    element=f"Width ({event_label})",
                    card="1..1",
                    datatype="DV_DURATION",
                    definition=self._event_width_definition(event_obj, event_label),
                    info=f"source=RM; path=data/events/width; event={event_label}; depth={depth}",
                )
            )
        return rows

    def _concept_cardinality(
        self,
        concept_obj: ET.Element,
        concept_attr: Optional[ET.Element],
        parent_object_occurrences: Optional[str],
    ) -> str:
        # Mapping rows should use the archetype concept cardinality (e.g. ELEMENT[at0008] occurrences),
        # not the cardinality of the internal DV_* value node.
        return (
            self._cardinality_from_object(concept_obj)
            or (self._cardinality_from_attribute(concept_attr) if concept_attr is not None else None)
            or parent_object_occurrences
            or "0..1"
        )

    def _display_cluster(self, term: ParsedTerm, path: str, internal_cluster_context: Optional[str] = None) -> str:
        label = term.text or self._prettify_segment(path.split("/")[-1])
        return self._apply_display_context(label, path, internal_cluster_context)

    def _display_leaf(
        self,
        path: str,
        term: ParsedTerm,
        event_label: Optional[str],
        internal_cluster_context: Optional[str] = None,
    ) -> str:
        if path.endswith("/time") and event_label:
            label = f"Time ({event_label})"
            return self._apply_display_context(label, path, internal_cluster_context)
        if path.endswith("/math_function") and event_label:
            label = f"Math function ({event_label})"
            return self._apply_display_context(label, path, internal_cluster_context)
        if path.endswith("/width") and event_label:
            label = f"Width ({event_label})"
            return self._apply_display_context(label, path, internal_cluster_context)

        label = term.text or self._prettify_segment(path.split("/")[-2] if path.endswith("/value") else path.split("/")[-1])
        return self._apply_display_context(label, path, internal_cluster_context)

    def _apply_display_context(self, label: str, path: str, internal_cluster_context: Optional[str] = None) -> str:
        contexts: List[str] = []
        suffix = self._context_suffix(path)
        if suffix:
            contexts.append(suffix)
        if internal_cluster_context and internal_cluster_context != label:
            contexts.append(internal_cluster_context)
        if contexts:
            return f"{label} ({' | '.join(dict.fromkeys(contexts))})"
        return label

    def _context_suffix(self, path: str) -> str:
        segments = [
            re.sub(r"\[[^\]]+\]", "", segment)
            for segment in path.strip("/").split("/")
            if segment
        ]

        if "protocol" in segments:
            return "protocol"

        if "events" in segments:
            if "state" in segments:
                return "events/state"
            if "data" in segments:
                return "events/data"
            return "events"

        if "activities" in segments and "description" in segments:
            return "activities/description"

        if "description" in segments:
            return "description"

        return ""

    def _has_internal_cluster_content(self, element: ET.Element) -> bool:
        for attr in self._children_named(element, "attributes"):
            for child in self._child_constraint_objects(attr):
                child_rm_type = (self._rm_type_name(child) or self._primitive_type_name(child) or "").upper()
                child_xsi_type = (self._xsi_type(child) or "").upper()
                if child_rm_type == "ELEMENT":
                    return True
                if child_rm_type == "CLUSTER" and child_xsi_type != "ARCHETYPE_SLOT":
                    return True
        return False

    def _is_cluster_slot(self, element: ET.Element) -> bool:
        xsi = (self._xsi_type(element) or "").upper()
        rm_type = (self._rm_type_name(element) or "").upper()
        has_includes = bool(self._children_named(element, "includes"))
        has_excludes = bool(self._children_named(element, "excludes"))

        if xsi == "ARCHETYPE_SLOT":
            return True

        if rm_type not in {"", "CLUSTER"}:
            return False

        has_slot_constraints = has_includes or has_excludes
        has_internal_children = self._has_internal_cluster_content(element)

        return has_slot_constraints and not has_internal_children

    def _cluster_includes_text(self, cluster_obj: ET.Element) -> str:
        labels: List[str] = []

        for include in self._children_named(cluster_obj, "includes"):
            patterns: List[str] = []

            for pattern_elem in self._iter_by_localname(include, "pattern"):
                text = (pattern_elem.text or "").strip()
                if text:
                    patterns.append(text)

            if not patterns:
                string_expression = self._text_of_first(include, "string_expression")
                if string_expression:
                    match = re.search(r"\{/([^}]*)/\}", string_expression)
                    patterns.append(match.group(1) if match else string_expression)

            for pattern in patterns:
                for alternative in self._split_unescaped_regex_alternatives(pattern):
                    label = self._humanize_cluster_include(alternative)
                    if label:
                        labels.append(label)

        labels = list(dict.fromkeys(labels))
        return f" Includes: {'; '.join(labels)}" if labels else ""

    def _split_unescaped_regex_alternatives(self, pattern: str) -> List[str]:
        parts: List[str] = []
        buffer: List[str] = []
        escaped = False
        for char in pattern:
            if escaped:
                buffer.append(char)
                escaped = False
            elif char == "\\":
                buffer.append(char)
                escaped = True
            elif char == "|":
                part = "".join(buffer).strip()
                if part:
                    parts.append(part)
                buffer = []
            else:
                buffer.append(char)
        part = "".join(buffer).strip()
        if part:
            parts.append(part)
        return parts

    def _humanize_cluster_include(self, pattern: str) -> Optional[str]:
        pattern = pattern.strip()
        if not pattern:
            return None
        if pattern == ".*":
            return "any CLUSTER archetype"

        match = re.fullmatch(
            r"openEHR-EHR-CLUSTER\\.([A-Za-z0-9_]+)(?:\(-\[a-zA-Z0-9_\]\+\)\*)?\\.v(\d+)",
            pattern,
        )
        if match:
            return f"{match.group(1)}_v{match.group(2)} and specialisations"

        match = re.fullmatch(r"openEHR-EHR-CLUSTER\\.([A-Za-z0-9_]+)\\.v(\d+)", pattern)
        if match:
            return f"{match.group(1)}_v{match.group(2)}"

        return None

    def _enrich_definition(self, attr_name: str, definition: str, event_label: Optional[str]) -> str:
        text = self._clean_definition_text(definition)
        if attr_name == "width" and event_label and event_label.lower() == "24 hour average":
            return "PT24H"
        return text

    def _is_internal_definition(self, text: str) -> bool:
        return "@ internal @" in (text or "").lower()

    def _clean_definition_text(self, text: str) -> str:
        text = re.sub(r"\s+", " ", (text or "").strip())
        if not text or text == "*" or self._is_internal_definition(text):
            return ""
        return text

    def _record_debug_warning(self, message: str) -> None:
        message = (message or "").strip()
        if message and message not in self.debug_warnings:
            self.debug_warnings.append(message)

    @classmethod
    def _load_openehr_terminology(cls) -> Dict[str, str]:
        """Load the external openEHR terminology XML once and cache it for reuse."""
        if cls._OPENEHR_TERMINOLOGY_CACHE is not None:
            return cls._OPENEHR_TERMINOLOGY_CACHE

        try:
            response = requests.get(cls.OPENEHR_TERMINOLOGY_URL, timeout=10)
            response.raise_for_status()
            root = ET.fromstring(response.text)
            mapping: Dict[str, str] = {}

            for concept in root.iter():
                local = concept.tag.split("}", 1)[-1] if isinstance(concept.tag, str) else ""
                if local != "concept":
                    continue
                code = (concept.attrib.get("id") or "").strip()
                rubric = (concept.attrib.get("rubric") or "").strip()
                if code and rubric and code not in mapping:
                    mapping[code] = rubric

            cls._OPENEHR_TERMINOLOGY_CACHE = mapping
            cls._OPENEHR_TERMINOLOGY_ERROR = None if mapping else (
                f"Could not parse any concepts from openEHR terminology XML at {cls.OPENEHR_TERMINOLOGY_URL}"
            )
            return mapping
        except Exception as exc:
            cls._OPENEHR_TERMINOLOGY_CACHE = {}
            cls._OPENEHR_TERMINOLOGY_ERROR = (
                f"Could not load openEHR terminology XML from {cls.OPENEHR_TERMINOLOGY_URL}: {exc}"
            )
            return {}

    def _extract_terminology_id_from_defining_code(self, defining_code: ET.Element) -> str:
        for candidate in defining_code.iter():
            if self._local(candidate.tag) == "terminology_id":
                value = self._text_of_child(candidate, "value")
                if value:
                    return value.strip()
        return ""

    def _resolve_coded_option_label(self, code: str, terminology_id: str) -> str:
        raw_code = (code or "").strip()
        raw_terminology = (terminology_id or "").strip()
        if not raw_code:
            return raw_code

        if raw_terminology.lower() == "local" or not raw_terminology:
            term = self.terms.get(raw_code) or self.terms.get(raw_code.lower()) or self.terms.get(raw_code.upper())
            return term.text if term and term.text else raw_code

        if raw_terminology.lower() == "openehr":
            mapping = self._load_openehr_terminology()
            label = mapping.get(raw_code)
            if label:
                return label

            if self._OPENEHR_TERMINOLOGY_ERROR:
                self._record_debug_warning(self._OPENEHR_TERMINOLOGY_ERROR)
                return f"openehr::{raw_code} [unresolved: terminology lookup failed]"
            self._record_debug_warning(f"Could not resolve openEHR terminology code openehr::{raw_code}")
            return f"openehr::{raw_code} [unresolved]"

        return f"{raw_terminology}::{raw_code}"

    def _extract_coded_text_options(self, dv_obj: ET.Element) -> List[str]:
        defining_code = self._find_attr_child(dv_obj, "defining_code")
        if defining_code is None:
            return []

        terminology_id = self._extract_terminology_id_from_defining_code(defining_code)
        options: List[str] = []
        seen: set[str] = set()

        def add_option(raw: str) -> None:
            raw = (raw or "").strip()
            if not raw:
                return
            label = self._resolve_coded_option_label(raw, terminology_id)
            if label not in seen:
                seen.add(label)
                options.append(label)

        for candidate in defining_code.iter():
            local = self._local(candidate.tag)
            raw = (candidate.text or "").strip()
            if local in {"code_list", "code_string", "code"} and raw:
                add_option(raw)

        if options:
            return options

        constraint_codes: List[str] = []
        seen_constraint_codes: set[str] = set()
        for candidate in defining_code.iter():
            texts = []
            if candidate.text:
                texts.append(candidate.text)
            texts.extend(str(v) for v in candidate.attrib.values())
            for raw_text in texts:
                for match in re.findall(r"\b(ac\d+)\b", raw_text or "", flags=re.I):
                    normalized = match.lower()
                    if normalized not in seen_constraint_codes:
                        seen_constraint_codes.add(normalized)
                        constraint_codes.append(match)

        for code in constraint_codes:
            term = self.terms.get(code) or self.terms.get(code.lower()) or self.terms.get(code.upper())
            if term:
                label = self._definition_text(term) or term.text
                if label and label not in seen:
                    seen.add(label)
                    options.append(label)
            elif code not in seen:
                seen.add(code)
                options.append(code)

        return options

    def _append_coded_text_options(self, definition: str, dv_obj: ET.Element, attr_name: str) -> str:
        options = self._extract_coded_text_options(dv_obj)
        if not options:
            return definition

        qualifier = "name in" if (attr_name or "").strip().lower() == "name" else "value in"
        options_text = "; ".join(options)
        suffix = f"DV_CODED_TEXT ({qualifier}): {options_text}"

        if definition:
            if suffix in definition:
                return definition
            return f"{definition} {suffix}".strip()
        return suffix

    def _definition_text(self, term: ParsedTerm) -> str:
        pieces = []
        for raw in (term.description, term.comment):
            cleaned = self._clean_definition_text(raw)
            if cleaned:
                pieces.append(cleaned)
        if not pieces:
            cleaned_text = self._clean_definition_text(term.text)
            if cleaned_text:
                pieces.append(cleaned_text)
        return " ".join(p for p in pieces if p)

    def _parse_info(self, info: str) -> Dict[str, str]:
        parsed: Dict[str, str] = {}
        for part in (info or "").split(";"):
            if "=" not in part:
                continue
            key, value = part.split("=", 1)
            key = key.strip()
            value = value.strip()
            if key:
                parsed[key] = value
        return parsed

    def _visible_path(self, path: str) -> str:
        return re.sub(r"/+", "/", path).strip("/")

    def _dedupe_rows(self, rows: List[MappingRow]) -> List[MappingRow]:
        seen = set()
        out: List[MappingRow] = []
        for row in rows:
            if not row.element:
                continue
            info = self._parse_info(row.info)
            key = (
                row.element,
                row.card,
                row.datatype,
                row.binding,
                row.definition,
                info.get("path", ""),
                info.get("node_id", ""),
            )
            if key in seen:
                continue
            seen.add(key)
            out.append(row)
        return out

    def _extract_archetype_id(self) -> str:
        for tag in ("archetype_id", "archetype-id", "archetypeid", "resource_id", "resource-id"):
            value = self._text_of_first(self.root, tag)
            if value and value.lower().startswith("openehr-"):
                return value
        for candidate in self.root.iter():
            text = (candidate.text or "").strip()
            if text.lower().startswith("openehr-"):
                return text
            for attr_val in candidate.attrib.values():
                if isinstance(attr_val, str) and attr_val.lower().startswith("openehr-"):
                    return attr_val
        return "archetype"

    def _extract_terms(self) -> Dict[str, ParsedTerm]:
        term_defs = list(self._iter_by_localname(self.root, "term_definitions"))
        chosen = None
        for lang in (self.preferred_language, "en"):
            for term_def in term_defs:
                if term_def.attrib.get("language", "").lower() == lang.lower():
                    chosen = term_def
                    break
            if chosen is not None:
                break
        if chosen is None and term_defs:
            chosen = term_defs[0]

        terms: Dict[str, ParsedTerm] = {}
        if chosen is None:
            return terms

        for item in self._children_named(chosen, "items"):
            code = item.attrib.get("code") or self._text_of_first(item, "code")
            if not code:
                continue

            fields: Dict[str, str] = {}
            for child in self._children_named(item, "items"):
                field_id = child.attrib.get("id", "").lower()
                value = (child.text or "").strip()
                if field_id and value:
                    fields[field_id] = value

            terms[code] = ParsedTerm(
                text=fields.get("text", ""),
                description=fields.get("description", ""),
                comment=fields.get("comment", ""),
            )
        return terms

    def _extract_bindings(self) -> Dict[str, str]:
        bindings: Dict[str, str] = {}
        for term_binding in self._iter_by_localname(self.root, "term_bindings"):
            terminology = term_binding.attrib.get("terminology", "") or self._text_of_first(term_binding, "terminology") or ""
            for item in self._children_named(term_binding, "items"):
                code = item.attrib.get("code") or self._text_of_first(item, "code")
                if not code:
                    continue
                value = self._find_first(item, "value")
                if value is None:
                    continue
                terminology_value = self._text_of_first(value, "value") or ""
                code_string = self._text_of_first(value, "code_string") or ""
                pieces = []
                if terminology or terminology_value:
                    pieces.append(terminology_value or terminology)
                if code_string:
                    pieces.append(code_string)
                binding_value = ": ".join(pieces) if pieces else ""
                if binding_value:
                    bindings[code] = binding_value
        return bindings

    def _cardinality_from_attribute(self, attr: ET.Element) -> Optional[str]:
        existence = self._first_child(attr, "existence")
        if existence is not None:
            return self._interval_to_cardinality(existence)
        cardinality = self._first_child(attr, "cardinality")
        if cardinality is not None:
            interval = self._first_child(cardinality, "interval")
            return self._interval_to_cardinality(interval or cardinality)
        return None

    def _cardinality_from_object(self, obj: ET.Element) -> Optional[str]:
        occurrences = self._first_child(obj, "occurrences")
        if occurrences is not None:
            return self._interval_to_cardinality(occurrences)
        return None

    def _interval_to_cardinality(self, element: Optional[ET.Element]) -> Optional[str]:
        if element is None:
            return None
        lower = self._text_of_first(element, "lower")
        upper = self._text_of_first(element, "upper")
        upper_unbounded = self._text_of_first(element, "upper_unbounded")
        if lower is None and upper is None:
            lower = element.attrib.get("lower")
            upper = element.attrib.get("upper")
            upper_unbounded = element.attrib.get("upper_unbounded")
        if lower is None and upper is None and upper_unbounded is None:
            return None
        if str(upper_unbounded).lower() == "true":
            upper = "*"
        if upper in ("-1", None, ""):
            upper = "*"
        lower = lower if lower not in (None, "") else "0"
        return f"{lower}..{upper}"

    def _child_constraint_objects(self, attr: ET.Element) -> List[ET.Element]:
        out: List[ET.Element] = []
        for child in list(attr):
            if self._local(child.tag) != "children":
                continue
            if self._looks_like_constraint_object(child):
                out.append(child)
                continue
            for grandchild in list(child):
                if self._looks_like_constraint_object(grandchild):
                    out.append(grandchild)
        return out

    def _looks_like_constraint_object(self, element: ET.Element) -> bool:
        local = self._local(element.tag)
        if local.startswith("c_"):
            return True
        if self._xsi_type(element):
            return True
        if self._rm_type_name(element):
            return True
        if element.attrib.get("node_id"):
            return True
        if self._first_child(element, "rm_type_name") is not None:
            return True
        return False

    def _primitive_type_name(self, element: ET.Element) -> Optional[str]:
        xsi = self._xsi_type(element)
        primitive_map = {
            "C_STRING": "String",
            "C_INTEGER": "Integer",
            "C_REAL": "Real",
            "C_BOOLEAN": "Boolean",
            "C_DATE": "Date",
            "C_DATE_TIME": "DateTime",
            "C_TIME": "Time",
            "C_DURATION": "Duration",
            "C_TERMINOLOGY_CODE": "Code",
        }
        return primitive_map.get(xsi)

    def _rm_type_name(self, element: ET.Element) -> Optional[str]:
        return self._text_of_child(element, "rm_type_name") or element.attrib.get("rm_type_name")

    def _xsi_type(self, element: ET.Element) -> Optional[str]:
        for key, value in element.attrib.items():
            if key.endswith("type"):
                return value.split(":")[-1]
        local = self._local(element.tag)
        return local.upper() if local.startswith("c_") else None

    def _build_path(self, parent_path: str, attr_name: str, node_id: Optional[str]) -> str:
        base = f"{parent_path}/{attr_name}" if parent_path else attr_name
        if node_id and re.match(r"^(at|ac|id)\d+", node_id):
            return f"{base}[{node_id}]"
        return base

    def _first_child(self, element: ET.Element, local_name: str) -> Optional[ET.Element]:
        for child in list(element):
            if self._local(child.tag) == local_name:
                return child
        return None

    def _text_of_child(self, element: ET.Element, local_name: str) -> Optional[str]:
        found = self._first_child(element, local_name)
        if found is not None and found.text is not None:
            text = found.text.strip()
            return text or None
        return None

    def _find_first(self, element: ET.Element, local_name: str) -> Optional[ET.Element]:
        for candidate in element.iter():
            if self._local(candidate.tag) == local_name:
                return candidate
        return None

    def _text_of_first(self, element: ET.Element, local_name: str) -> Optional[str]:
        found = self._find_first(element, local_name)
        if found is not None and found.text is not None:
            text = found.text.strip()
            return text or None
        return None

    def _children_named(self, element: ET.Element, local_name: str) -> Iterable[ET.Element]:
        for child in list(element):
            if self._local(child.tag) == local_name:
                yield child

    def _iter_by_localname(self, element: ET.Element, local_name: str) -> Iterable[ET.Element]:
        for candidate in element.iter():
            if self._local(candidate.tag) == local_name:
                yield candidate

    def _prettify_segment(self, value: str) -> str:
        value = re.sub(r"\[.*?\]", "", value)
        value = value.replace("_", " ").strip()
        return value[:1].upper() + value[1:] if value else value

    @staticmethod
    def _local(tag: str) -> str:
        return tag.split("}")[-1].split(":")[-1]


# ---------------------------------------------------------------------------
# Optional template-based validation and legend copying
# ---------------------------------------------------------------------------
class DataValidationAdder:
    """Copy validation sheets from a template workbook into the generated output."""
    def __init__(self, template_path: str):
        self.template = openpyxl.load_workbook(pathlib.Path(template_path))

    def add_validation(self, workbook_path: pathlib.Path) -> None:
        """Apply template-driven list validation and legend sheets to the output workbook."""
        ensure_output_path_writable(workbook_path, "Excel output file")
        workbook = openpyxl.load_workbook(workbook_path)
        structure_sheet = workbook[workbook.sheetnames[0]]
        headers = {cell.value: cell.column_letter for cell in structure_sheet[1]}
        for header in headers:
            legend_name = "Legenda" + str(header).capitalize()
            if legend_name in self.template:
                self._copy_sheet(workbook, legend_name)
                self._add_data_validation(workbook, structure_sheet.title, legend_name, headers[header])
        ensure_output_path_writable(workbook_path, "Excel output file")
        workbook.save(workbook_path)

    def _copy_sheet(self, workbook: openpyxl.Workbook, sheet_name: str) -> None:
        new_sheet = workbook.create_sheet(sheet_name)
        for row in self.template[sheet_name].iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)

    def _add_data_validation(self, workbook: openpyxl.Workbook, target_sheet_name: str, legend_name: str, column: str) -> None:
        max_row = len(list(workbook[legend_name].rows))
        formula = f"={legend_name}!$A$2:$A${max_row + 1}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        workbook[target_sheet_name].add_data_validation(dv)
        dv.add(f"{column}2:{column}1000")

        for row in range(2, max_row + 3):
            cell = workbook[legend_name].cell(row, 1)
            if cell.fill.fgColor.rgb != "00000000":
                rule = openpyxl.formatting.rule.CellIsRule(
                    operator="equal",
                    formula=[f'{legend_name}!${cell.column_letter}${cell.row}'],
                    fill=openpyxl.styles.PatternFill(
                        start_color=cell.fill.fgColor.rgb,
                        end_color=cell.fill.fgColor.rgb,
                        fill_type="solid",
                    ),
                )
                workbook[target_sheet_name].conditional_formatting.add(f"{column}2:{column}1000", rule)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------
def main(argv: List[str]) -> int:
    """Parse command line arguments and run the conversion."""
    parser = argparse.ArgumentParser(
        description=(
            "Convert an openEHR CKM archetype XML into a mapping-friendly Excel overview. "
            "By default, fetch mode is auto: v0 archetypes use the latest development revision, "
            "and v1+ archetypes use the latest published revision."
        )
    )
    parser.add_argument(
        "archetype",
        help="Archetype id or CKM search text, for example openEHR-EHR-ADMIN_ENTRY.admission.v0",
    )
    parser.add_argument(
        "output",
        help="Output folder where the generated Excel and debug_archetype.xml will be written.",
    )
    parser.add_argument(
        "template",
        nargs="?",
        default=None,
        help="Optional Excel template used to copy validations after generation.",
    )
    parser.add_argument(
        "--fetch-mode",
        choices=["auto", "published", "development"],
        default="auto",
        help=(
            "CKM fetch mode. auto = v0 -> latest development revision, v1+ -> latest published revision; "
            "published = always request latest published revision; "
            "development = always request latest development revision."
        ),
    )

    args = parser.parse_args(argv[1:])

    converter = ArchetypeToSpreadsheet(output_folder=args.output)
    out_path = converter.convert(args.archetype, fetch_mode=args.fetch_mode)
    print(f"Created: {out_path}")

    if args.template:
        DataValidationAdder(args.template).add_validation(out_path)
        print(f"Added validations from template: {args.template}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
