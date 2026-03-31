import argparse
import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_OUTPUT_SHEET = "Generated"


def normalize_label(text: str) -> str:
    text = (text or "").strip().lower()
    text = text.replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def load_rows(xlsx_path: Path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[DEFAULT_OUTPUT_SHEET] if DEFAULT_OUTPUT_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    header_map = {name: idx for idx, name in enumerate(headers)}

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row]
        if all(v in (None, "") for v in values):
            continue

        name = str(values[header_map.get("Element", 0)] or values[0] or "")
        card = str(values[header_map.get("Card", 2)] or values[2] or "")
        row_type = str(values[header_map.get("Datatype", 3)] or values[3] or "")
        definition = str(values[header_map.get("Definition", 5)] or values[5] or "")
        info = str(values[header_map.get("Aanvullende informatie", 8)] or values[8] or "")

        leading_spaces = len(name) - len(name.lstrip(" "))
        excel_indent = row[header_map.get("Element", 0)].alignment.indent or 0

        rows.append({
            "row_idx": row_idx,
            "element": name,
            "element_stripped": name.strip(),
            "element_normalized": normalize_label(name),
            "card": card,
            "card_normalized": normalize_label(card),
            "type": row_type,
            "type_normalized": normalize_label(row_type),
            "definition": definition,
            "definition_normalized": definition.lower(),
            "info": info,
            "info_normalized": info.lower(),
            "indent_spaces": leading_spaces,
            "indent_excel": int(excel_indent),
        })
    return rows


def run_script(script: Path, archetype: str, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)
    cmd = [sys.executable, str(script), archetype, str(output_dir)]
    completed = subprocess.run(cmd, capture_output=True, text=True)
    if completed.returncode != 0:
        raise AssertionError(
            f"Script failed for {archetype}\n"
            f"STDOUT:\n{completed.stdout}\n\nSTDERR:\n{completed.stderr}"
        )
    return completed


def indent_value(row):
    return max(row["indent_spaces"] // 4, row["indent_excel"])


def row_matches_element(row, spec) -> bool:
    target = normalize_label(spec.get("element", ""))
    contains_target = normalize_label(spec.get("element_contains", ""))

    actual = row["element_normalized"]
    if target:
        if actual == target or target in actual or actual in target:
            return True
        return False
    if contains_target:
        return contains_target in actual
    return True


def row_matches_type(row, spec) -> bool:
    if "type" in spec and row["type"] != spec["type"]:
        return False
    if "type_contains" in spec and normalize_label(spec["type_contains"]) not in row["type_normalized"]:
        return False
    return True


def row_matches_card(row, spec) -> bool:
    if "card" in spec and row["card"] != spec["card"]:
        return False
    if "card_contains" in spec and normalize_label(spec["card_contains"]) not in row["card_normalized"]:
        return False
    return True


def find_row(rows, spec):
    candidates = []
    for row in rows:
        if row_matches_element(row, spec) and row_matches_type(row, spec) and row_matches_card(row, spec):
            candidates.append(row)

    target = normalize_label(spec.get("element", ""))
    if target:
        for row in candidates:
            if row["element_normalized"] == target:
                return row
    return candidates[0] if candidates else None


def find_rows_contains(rows, text):
    needle = normalize_label(text)
    return [row for row in rows if needle in row["element_normalized"]]


def subtree_rows(rows, parent):
    parent_indent = indent_value(parent)
    result = []

    for row in rows:
        if row["row_idx"] <= parent["row_idx"]:
            continue
        current_indent = indent_value(row)
        if current_indent <= parent_indent:
            break
        result.append(row)

    return result


def assert_contains(rows, spec):
    row = find_row(rows, spec)
    assert row is not None, f"Missing row: {spec}"

    if "definition_contains" in spec:
        assert spec["definition_contains"].lower() in row["definition_normalized"], (
            f"Definition for {row['element_stripped']} does not contain {spec['definition_contains']!r}. "
            f"Actual: {row['definition']!r}"
        )

    if "definition_contains_any" in spec:
        assert any(s.lower() in row["definition_normalized"] for s in spec["definition_contains_any"]), (
            f"Definition for {row['element_stripped']} did not contain any of {spec['definition_contains_any']}. "
            f"Actual: {row['definition']!r}"
        )

    if "definition_contains_all" in spec:
        missing = [s for s in spec["definition_contains_all"] if s.lower() not in row["definition_normalized"]]
        assert not missing, (
            f"Definition for {row['element_stripped']} missed {missing}. Actual: {row['definition']!r}"
        )

    if "info_contains" in spec:
        assert spec["info_contains"].lower() in row["info_normalized"], (
            f"Info for {row['element_stripped']} does not contain {spec['info_contains']!r}. "
            f"Actual: {row['info']!r}"
        )


def assert_indented_under(rows, spec):
    parent = find_row(rows, {"element": spec["parent"]})
    assert parent is not None, f"Missing parent row: {spec['parent']}"

    subrows = subtree_rows(rows, parent)
    assert subrows, f"Parent {parent['element_stripped']} has no subtree rows"

    child = None
    for row in subrows:
        if normalize_label(spec["child_contains"]) in row["element_normalized"]:
            child = row
            break

    assert child is not None, (
        f"Missing child containing {spec['child_contains']!r} within subtree of {parent['element_stripped']}. "
        f"Subtree rows: {[row['element_stripped'] for row in subrows[:20]]}"
    )

    assert child["row_idx"] > parent["row_idx"], (
        f"Child {child['element_stripped']} is not below parent {parent['element_stripped']}"
    )
    assert indent_value(child) > indent_value(parent), (
        f"Child {child['element_stripped']} is not indented below parent {parent['element_stripped']}. "
        f"Parent indent={indent_value(parent)}, child indent={indent_value(child)}"
    )


def assert_absent(rows, spec):
    matches = find_rows_contains(rows, spec["element_contains"])
    assert not matches, (
        f"Expected no rows containing {spec['element_contains']!r}, but found: "
        f"{[row['element_stripped'] for row in matches]}"
    )


def main():
    parser = argparse.ArgumentParser(description="Run semantic regression tests for the archetype spreadsheet generator.")
    parser.add_argument("script", help="Path to ArchetypeToSpreadsheet version to test")
    parser.add_argument("--expectations", default="regression_expectations_v5.json", help="JSON file with semantic checks")
    parser.add_argument("--workdir", default="regression_output", help="Output folder for generated files")
    args = parser.parse_args()

    script = Path(args.script).resolve()
    expectations_path = Path(args.expectations).resolve()
    workdir = Path(args.workdir).resolve()

    expectations = json.loads(expectations_path.read_text(encoding="utf-8"))

    failures = []
    for case_name, case in expectations.items():
        case_dir = workdir / case_name
        try:
            run_script(script, case["archetype"], case_dir)
            xlsx_candidates = sorted(case_dir.glob("*.xlsx"))
            assert xlsx_candidates, f"No xlsx generated in {case_dir}"
            rows = load_rows(xlsx_candidates[0])

            for spec in case.get("contains", []):
                assert_contains(rows, spec)
            for spec in case.get("indented_under", []):
                assert_indented_under(rows, spec)
            for spec in case.get("absent", []):
                assert_absent(rows, spec)

            print(f"[PASS] {case_name}")
        except Exception as exc:
            failures.append((case_name, str(exc)))
            print(f"[FAIL] {case_name}: {exc}")

    if failures:
        print("\nFailures:")
        for case_name, msg in failures:
            print(f"- {case_name}: {msg}")
        raise SystemExit(1)

    print("\nAll regression checks passed.")


if __name__ == "__main__":
    main()