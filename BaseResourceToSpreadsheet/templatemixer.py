import copy
import openpyxl
import pathlib

from openpyxl.worksheet.datavalidation import DataValidation

class TemplateMixer:
    def __init__(self, template_path):
        self.template = openpyxl.load_workbook(pathlib.Path(template_path))

        self.mapping_tables = {}
        for sheet in self.template:
            if sheet.title.startswith("Mapping"):
                mapping_table = {row[0].value: row[1].value for row in sheet.iter_rows(min_row = 2)}
                self.mapping_tables[sheet.title] = mapping_table

    def mix(self, workbook_path):
        workbook = openpyxl.load_workbook(workbook_path)
        headers = {cell.column_letter: cell.value.split(":")[0] for cell in workbook["Structure"][1]}
        for col in headers:
            legend_name = "Legenda" + headers[col]
            mapping_name = "Mapping" + headers[col]
            if legend_name in self.template:
                self.__copySheet__(workbook, legend_name)
                self.__addDataValidation__(workbook, legend_name, col)
            if mapping_name in self.template:
                self.__map__(workbook, mapping_name, col)

        workbook.save(workbook_path)

    def __copySheet__(self, workbook, sheet_name):
        if sheet_name in workbook:
            # Don't add it if it already exists
            return

        new_sheet = workbook.create_sheet(sheet_name)
        for row in self.template[sheet_name].iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value

                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)

    def __addDataValidation__(self, workbook, sheet_name, column):
        max_row = len(list(workbook[sheet_name].rows))

        # Assume the first row is the header row, and data is in the first col
        formula = f"={sheet_name}!$A$2:$A${max_row + 1}"
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True
        )
        workbook["Structure"].add_data_validation(dv)
        dv.add(f"{column}2:{column}1000")

        # Also add conditional formatting based on the colors in the source sheet
        default_style = workbook[sheet_name].cell(1, 3).fill
        for row in range(2, max_row + 3):
            cell = workbook[sheet_name].cell(row, 1)
            if (cell.fill.fgColor.rgb != "00000000"):
                rule = openpyxl.formatting.rule.CellIsRule(
                    operator="equal",
                    formula=[f'{sheet_name}!${cell.column_letter}${cell.row}'],
                    fill=openpyxl.styles.PatternFill(start_color=cell.fill.fgColor.rgb, end_color=cell.fill.fgColor.rgb, fill_type="solid")
                )
                workbook["Structure"].conditional_formatting.add(f"{column}2:{column}1000", rule)

    def __map__(self, workbook, sheet_name, col):
        mapping_table = self.mapping_tables[sheet_name]
        
        for cell in workbook["Structure"][col][2:]:
            values = [value.strip() for value in cell.value.split(",")]
            new_values = []
            for value in values:
                parts = value.split("(")
                if parts[0] in mapping_table:
                    new_type = mapping_table[parts[0]]
                    new_values.append("(".join([new_type] + parts[1:]))
                else:
                    print(f"Unmappable: {value}")
                    new_values.append(value)
            cell.value = ", ".join(new_values)